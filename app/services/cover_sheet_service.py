"""
Service for generating and printing cover sheets.
"""

import os
import subprocess
import platform
import tempfile
from datetime import datetime
from typing import List, Optional
from pathlib import Path

from openpyxl import load_workbook

from ..utils.logger import logger, log_exception


class CoverSheetService:
    """Handles cover sheet generation and printing."""

    # Cell mappings for the cover sheet template
    # Format: (row, col) - 1-indexed for openpyxl
    CELL_MAP = {
        "kamp": (1, 1),        # A1 - KaMP#
        "date": (1, 9),        # I1 - Date
        "operator": (4, 4),    # D4 - Operator Name
        "climate": (5, 4),     # D5 - Climate
        "event_name": (6, 4),  # D6 - Event Name (error type)
        "unit": (7, 4),        # D7 - Unit#
        "media": (8, 4),       # D8 - Media#
        "script": (9, 4),      # D9 - Script#
        "plexity": (10, 4),    # D10 - Plexity
        "load": (11, 4),       # D11 - Load
        "run": (12, 4),        # D12 - Run
        "pages": (13, 4),      # D13 - Page in Run
        "phase": (14, 4),      # D14 - Phase
        "firmware": (15, 4),   # D15 - Firmware
    }

    def __init__(self, template_path: str, output_dir: Optional[str] = None, keep_files: bool = True):
        """
        Initialize the cover sheet service.

        Args:
            template_path: Path to the cover sheet template Excel file
            output_dir: Directory to save generated cover sheets (default: temp dir)
            keep_files: If True, keep generated files after printing
        """
        self.template_path = template_path
        self.keep_files = keep_files

        if output_dir:
            self.output_dir = output_dir
            os.makedirs(self.output_dir, exist_ok=True)
        else:
            # Use temp directory if not keeping files
            self.output_dir = tempfile.gettempdir()

    def _fill_template(
        self,
        error_type: str,
        printer: str,
        pages: List[int],
        kamp: str = "",
        operator: str = "",
        climate: str = "",
        media: str = "",
        script: str = "",
        plexity: str = "",
        load: str = "",
        run: str = "",
        phase: str = "",
        firmware: str = "",
        date: Optional[datetime] = None
    ) -> str:
        """
        Fill the template and save to a temp file.

        Returns:
            Path to the generated cover sheet file
        """
        # Load the template
        wb = load_workbook(self.template_path)
        ws = wb.active

        # Fill in the values
        date = date or datetime.now()
        pages_str = ", ".join(map(str, pages))

        # Helper to set cell value (row/col are now 1-indexed for openpyxl)
        def set_cell(key: str, value):
            row, col = self.CELL_MAP[key]
            ws.cell(row=row, column=col, value=value)

        set_cell("kamp", f"KaMP# {kamp}" if kamp else "KaMP#")
        set_cell("date", date)
        set_cell("operator", operator)
        set_cell("climate", climate)
        set_cell("event_name", error_type)
        set_cell("unit", printer)
        set_cell("media", media)
        set_cell("script", script)
        set_cell("plexity", plexity)
        set_cell("load", load)
        set_cell("run", run)
        set_cell("pages", pages_str)
        set_cell("phase", phase)
        set_cell("firmware", firmware)

        # Set print area explicitly
        ws.print_area = 'A1:I15'

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        safe_error = error_type.replace(" ", "_").replace("/", "-")[:30]
        filename = f"CoverSheet_{printer[-3:]}_{safe_error}_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, filename)

        # Save
        wb.save(output_path)
        wb.close()

        return output_path

    def _print_via_excel_mac(self, file_path: str) -> bool:
        """Print using Microsoft Excel on macOS via AppleScript."""
        abs_path = os.path.abspath(file_path)
        logger.debug(f"Attempting Excel print on macOS: {abs_path}")
        
        # Use System Events to simulate Cmd+P for more reliable printing
        script = f'''
        tell application "Microsoft Excel"
            activate
            open "{abs_path}"
            delay 1
        end tell

        tell application "System Events"
            tell process "Microsoft Excel"
                keystroke "p" using command down
                delay 0.5
                keystroke return
            end tell
        end tell

        delay 2

        tell application "Microsoft Excel"
            close active workbook saving no
        end tell
        '''
        try:
            result = subprocess.run(
                ["osascript", "-e", script],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0:
                logger.info(f"Printed via Excel: {os.path.basename(file_path)}")
                return True
            else:
                logger.warning(f"Excel print failed: {result.stderr}")
                return False
        except subprocess.TimeoutExpired:
            logger.error("Excel print timed out after 30 seconds")
            return False
        except Exception as e:
            log_exception(e, "Excel print error")
            return False

    def _print_via_numbers_mac(self, file_path: str) -> bool:
        """Print using Numbers on macOS via AppleScript (fallback)."""
        abs_path = os.path.abspath(file_path)
        logger.debug(f"Attempting Numbers print on macOS: {abs_path}")
        
        script = f'''
        tell application "Numbers"
            open "{abs_path}"
            print front document
            close front document saving no
        end tell
        '''
        try:
            result = subprocess.run(
                ["osascript", "-e", script],
                capture_output=True,
                text=True,
                timeout=30
            )
            if result.returncode == 0:
                logger.info(f"Printed via Numbers: {os.path.basename(file_path)}")
                return True
            else:
                logger.warning(f"Numbers print failed: {result.stderr}")
                return False
        except Exception as e:
            log_exception(e, "Numbers print error")
            return False

    def _print_via_win32(self, file_path: str) -> bool:
        """Print using Excel COM automation on Windows."""
        logger.debug(f"Attempting Excel COM print on Windows: {file_path}")
        
        try:
            import win32com.client

            # Try to connect to existing Excel instance, or create new one
            try:
                excel = win32com.client.GetActiveObject("Excel.Application")
                excel_was_running = True
                logger.debug("Connected to existing Excel instance")
            except:
                excel = win32com.client.Dispatch("Excel.Application")
                excel_was_running = False
                logger.debug("Started new Excel instance")

            abs_path = os.path.abspath(file_path)
            wb = excel.Workbooks.Open(abs_path)
            wb.PrintOut()
            wb.Close(SaveChanges=False)

            # Only quit Excel if we started it
            if not excel_was_running:
                excel.Quit()
                logger.debug("Closed Excel instance")

            logger.info(f"Printed via Excel COM: {os.path.basename(file_path)}")
            return True
        except ImportError:
            logger.error("win32com not installed. Install with: pip install pywin32")
            return False
        except Exception as e:
            log_exception(e, "Excel COM print error")
            return False

    def print_file(self, file_path: str, delete_after: bool = False) -> bool:
        """
        Print a file using the system's Excel application.

        Args:
            file_path: Path to the Excel file to print
            delete_after: If True, delete the file after printing

        Returns:
            True if print command was sent successfully
        """
        if not os.path.exists(file_path):
            logger.error(f"File not found: {file_path}")
            return False

        success = False
        system = platform.system()
        logger.debug(f"Printing on {system}: {file_path}")

        if system == "Darwin":  # macOS
            # Try Excel first, fall back to Numbers
            success = self._print_via_excel_mac(file_path)
            if not success:
                logger.info("Trying Numbers as fallback...")
                success = self._print_via_numbers_mac(file_path)

        elif system == "Windows":
            success = self._print_via_win32(file_path)

        else:  # Linux - try lpr with libreoffice conversion
            logger.warning(f"Linux printing not fully supported. File saved at: {file_path}")
            success = False

        # Clean up temp file if requested
        if delete_after and success and not self.keep_files:
            try:
                os.remove(file_path)
                logger.debug(f"Deleted temp file: {file_path}")
            except Exception as e:
                logger.warning(f"Could not delete temp file: {e}")

        return success

    def print_cover_sheet(
        self,
        error_type: str,
        printer: str,
        pages: List[int],
        kamp: str = "",
        operator: str = "",
        climate: str = "",
        media: str = "",
        script: str = "",
        plexity: str = "",
        load: str = "",
        run: str = "",
        phase: str = "",
        firmware: str = "",
    ) -> bool:
        """
        Generate and print a single cover sheet directly.

        Returns:
            True if printed successfully
        """
        file_path = self._fill_template(
            error_type=error_type,
            printer=printer,
            pages=pages,
            kamp=kamp,
            operator=operator,
            climate=climate,
            media=media,
            script=script,
            plexity=plexity,
            load=load,
            run=run,
            phase=phase,
            firmware=firmware,
        )

        return self.print_file(file_path, delete_after=not self.keep_files)

    def print_from_entry(self, entry) -> int:
        """
        Print cover sheets directly from an ErrorEntry.

        Args:
            entry: ErrorEntry object with all required fields

        Returns:
            Number of sheets successfully printed
        """
        printed = 0

        for printer in entry.printers:
            success = self.print_cover_sheet(
                error_type=entry.error_type,
                printer=printer,
                pages=entry.pages,
                kamp=entry.kamp,
                operator=entry.operator,
                climate=entry.climate,
                media=entry.media,
                script=entry.script,
                plexity=entry.plexity,
                load=entry.load,
                run=entry.run,
                phase=entry.phase,
                firmware=entry.firmware,
            )
            if success:
                printed += 1

        return printed

    def save_from_entry(self, entry, output_dir: str) -> List[str]:
        """
        Generate and save cover sheets from an ErrorEntry WITHOUT printing.

        Args:
            entry: ErrorEntry object with all required fields
            output_dir: Directory to save the cover sheet files

        Returns:
            List of paths to the saved cover sheet files
        """
        saved_files = []
        
        # Temporarily override output_dir
        original_output_dir = self.output_dir
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        for printer in entry.printers:
            file_path = self._fill_template(
                error_type=entry.error_type,
                printer=printer,
                pages=entry.pages,
                kamp=entry.kamp,
                operator=entry.operator,
                climate=entry.climate,
                media=entry.media,
                script=entry.script,
                plexity=entry.plexity,
                load=entry.load,
                run=entry.run,
                phase=entry.phase,
                firmware=entry.firmware,
            )
            saved_files.append(file_path)

        # Restore original output_dir
        self.output_dir = original_output_dir
        
        return saved_files

    # Legacy methods for backwards compatibility
    def generate_cover_sheet(self, **kwargs) -> str:
        """Generate a cover sheet file (legacy method)."""
        return self._fill_template(**kwargs)

    def generate_from_entry(self, entry, test_info: dict) -> List[str]:
        """Generate cover sheet files from an ErrorEntry (legacy method)."""
        generated_files = []

        for printer in entry.printers:
            path = self._fill_template(
                error_type=entry.error_type,
                printer=printer,
                pages=entry.pages,
                kamp=entry.kamp,
                operator=entry.operator,
                climate=test_info.get("climate", ""),
                media=test_info.get("media", ""),
                script=test_info.get("script", ""),
                plexity=test_info.get("plexity", ""),
                load=test_info.get("load", ""),
                run=test_info.get("run", ""),
                phase=test_info.get("phase", ""),
                firmware=entry.firmware,
            )
            generated_files.append(path)

        return generated_files
